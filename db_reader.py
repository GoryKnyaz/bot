from zipfile import ZipFile

from openpyxl import load_workbook
from xlwt import Workbook


def write_to_excel(filename, headlines, mass):
    """
    A function to write data to excel table.
    :param str filename: the name of the file by which the Excel file will be stored.
    :param List(str) headlines: the array of names of columns.
    :param List(List(Any)) mass: two-dimensional array of data.
    """
    book = Workbook(encoding='utf8')
    sh = book.add_sheet('Table0')

    for headline_index, headline in enumerate(headlines):
        sh.write(0, headline_index, headline)

    for index_row, value_elements in enumerate(mass, start=1):
        for index_col, value_element in enumerate(mass[index_row - 1]):
            sh.write(index_row, index_col, value_element)

    # Automatic size of columns based on their length
    col_width_mass = [len(headline) for headline in headlines]
    for row_of_mass in mass:
        for index_col_mass, col_of_row in enumerate(row_of_mass):
            if len(str(col_of_row)) > col_width_mass[index_col_mass]:
                col_width_mass[index_col_mass] = len(col_of_row)

    for index, col_width in enumerate(col_width_mass):
        sh.col(index).width = 256 * (col_width + 1)
    book.save(filename)


class Reader:
    """
    This is a class designed to read a certain type of file.
    (currently only: txt, xlsx, txt found in zip) and store its contents.
    """

    def __init__(self, path_to_file=''):
        """
        Constructor.
        :param str path_to_file: path to the file for reading and storing information in itself.
        """
        self.data = None
        self.path_to_file = None
        if not path_to_file:
            self.read(path_to_file)

    def read(self, path_to_file):
        """
        A function to read information from a file and write it to an array called 'data'.
        :param str path_to_file: path to the file for reading and storing information in itself.
        """
        self.path_to_file = path_to_file
        self.data = []
        if path_to_file.endswith('txt'):
            with open(path_to_file, encoding="utf8") as txt_file:
                self.data = [data for data in txt_file.readlines()]
        elif path_to_file.endswith('xlsx'):
            workbook = load_workbook(path_to_file)
            worksheet = workbook.active
            for col in worksheet.iter_cols(1, worksheet.max_column):
                for row in range(0, worksheet.max_row):
                    if col[row].value is not None:
                        self.data.append(col[row].value)
        elif path_to_file.endswith('zip'):
            with ZipFile(path_to_file) as archive:
                for archive_filename in archive.namelist():
                    if archive_filename.endswith('txt'):
                        with archive.open(archive_filename) as txt_file:
                            self.data += [data.decode() for data in txt_file.readlines()]
